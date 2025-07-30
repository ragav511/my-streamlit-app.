# File: complete_boq_po_system.py
import streamlit as st
import pandas as pd
from utils.dual_db import get_connection, db_manager, backup_now, get_backup_status, test_server_connection
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
import io
from io import BytesIO
import os
from decimal import Decimal
from dotenv import load_dotenv
from PIL import Image
from num2words import num2words
import re
import sqlite3
import bcrypt
from sqlalchemy import create_engine, text
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Initialize session state for authentication
if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False
    st.session_state['role'] = None
    st.session_state['user_id'] = None
    st.session_state['username'] = None

# Set page configuration
st.set_page_config(page_title="BOQ & PO Management System", layout="wide")

# Load environment variables
load_dotenv()

# Database setup with both PostgreSQL and SQLite support
def init_sqlite_db():
    """Initialize SQLite database for authentication"""
    engine = create_engine('sqlite:///boq_po_auth.db', connect_args={'check_same_thread': False})
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password_hash TEXT,
                role TEXT,
                name TEXT,
                email TEXT,
                contact_number TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        # Create default admin user
        hashed = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt())
        conn.execute(text("INSERT OR IGNORE INTO users (username, password_hash, role, name) VALUES (:username, :password_hash, :role, :name)"),
                     {'username': 'admin', 'password_hash': hashed, 'role': 'admin', 'name': 'Administrator'})
        conn.commit()
    return engine

# Initialize SQLite for auth
auth_engine = init_sqlite_db()

# Get PostgreSQL connection for main data
conn = get_connection()
cursor = conn.cursor()

# Authentication functions
def login_page():
    st.title("🔒 BOQ & PO Management System Login")
    
    # Login form
    with st.form("login_form"):
        st.subheader("Please log in to continue")
        username = st.text_input("Username", placeholder="Enter your username")
        password = st.text_input("Password", type="password", placeholder="Enter your password")
        submit = st.form_submit_button("🚀 Login", use_container_width=True)
        
        if submit:
            with auth_engine.connect() as conn:
                result = conn.execute(text("SELECT * FROM users WHERE username = :username"),
                                     {'username': username})
                user = result.mappings().fetchone()
                if user and bcrypt.checkpw(password.encode('utf-8'), user['password_hash']):
                    st.session_state['logged_in'] = True
                    st.session_state['role'] = user['role']
                    st.session_state['user_id'] = user['id']
                    st.session_state['username'] = user['username']
                    st.session_state['user_name'] = user['name']
                    st.success("✅ Logged in successfully!")
                    st.rerun()
                else:
                    st.error("❌ Invalid username or password")

def logout():
    st.session_state['logged_in'] = False
    st.session_state['role'] = None
    st.session_state['user_id'] = None
    st.session_state['username'] = None
    st.session_state['user_name'] = None
    st.success("✅ Logged out successfully!")
    st.rerun()

# Database table creation functions
def create_suppliers_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            address TEXT,
            gst_number VARCHAR(50),
            contact_person VARCHAR(255),
            contact_number VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

def create_bill_to_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bill_to_companies (
            id SERIAL PRIMARY KEY,
            company_name VARCHAR(255) NOT NULL,
            address TEXT,
            gst_number VARCHAR(50),
            contact_person VARCHAR(255),
            contact_number VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

def create_ship_to_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ship_to_addresses (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            address TEXT,
            gst_number VARCHAR(50),
            contact_person VARCHAR(255),
            contact_number VARCHAR(255),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

def create_locations_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS locations (
            id SERIAL PRIMARY KEY,
            location_code VARCHAR(10) NOT NULL UNIQUE,
            location_name VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

def create_po_counters_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS po_counters (
            id SERIAL PRIMARY KEY,
            location_code VARCHAR(10) NOT NULL,
            last_serial_number INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(location_code)
        )
    """)
    conn.commit()

def create_projects_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by INTEGER REFERENCES users(id)
        )
    """)
    conn.commit()

def create_boq_items_table():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS boq_items (
            id SERIAL PRIMARY KEY,
            project_id INTEGER REFERENCES projects(id) ON DELETE CASCADE,
            boq_ref VARCHAR(100),
            description TEXT,
            make VARCHAR(255),
            model VARCHAR(255),
            unit VARCHAR(50),
            boq_qty DECIMAL(12,2) DEFAULT 0,
            rate DECIMAL(12,2) DEFAULT 0,
            amount DECIMAL(12,2) DEFAULT 0,
            delivered_qty_1 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_2 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_3 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_4 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_5 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_6 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_7 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_8 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_9 DECIMAL(12,2) DEFAULT 0,
            delivered_qty_10 DECIMAL(12,2) DEFAULT 0,
            total_delivery_qty DECIMAL(12,2) DEFAULT 0,
            balance_to_deliver DECIMAL(12,2) DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

# Initialization functions with predefined data
def initialize_suppliers():
    cursor.execute("SELECT COUNT(*) FROM suppliers")
    count = cursor.fetchone()[0]
    
    if count == 0:
        suppliers_data = [
            ("SRK Trading Company", "2nd Floor, Front Side, Gali No- 23, 357/A, Khasra No- 643/2, Pandit Muhalla, Mandawali", "", "Mr. Rajkumar, Anuj", "91-9560114107, 91-9810529008"),
            ("4 Squares Corporation", "15TH MAIN, HBR LAYOUT BANGALORE, Karnataka - 560043", "29AAAFZ7997B1ZL", "Mr. Tijo George, Anuj Kumar", "91-9560114107, 9-97420 02555"),
            ("Wave Enterprise", "222, GALA HUB, Near Honest Restaurant, Gala Gymkhana Road, South Bopal, Ahmedabad - 380058", "24ECSPM2013M1ZD", "Anuj Kumar", "91-9560114107, 91-7874063796"),
            ("SPARK TECHNOLOGIES PRIVATE LIMITED", "56 Nehru Place, 308-309, Eros Apartment NEW DELHI Delhi - 110019", "07AAACS4609L1Z8", "Mr. Virat, Anuj Kumar", "91-9560114107, 91 98182 28701"),
            ("SUPERTRON ELECTRONICS PRIVATE LIMITED", "SUPERTRON HOUSE, 2 COOPER LANE, COOPER LANE, Kolkata, West Bengal, 700001", "07AADCS5971L1ZU", "Mr. Jagdish, Mr. Anuj Kumar", "91-9560114107, 91-9891289203"),
            ("SS INFOTECH", "First Floor Bajaj House 106A, 97 Nehru Place Flyover, Sona Sweets, New Delhi- 110019", "07GXPPS8415D1ZY", "Mr. Sahil, Mr. Anuj Kumar", "91-9560114107")
        ]
        
        for supplier in suppliers_data:
            cursor.execute("""
                INSERT INTO suppliers (name, address, gst_number, contact_person, contact_number)
                VALUES (%s, %s, %s, %s, %s)
            """, supplier)
        conn.commit()
        db_manager.backup_table('suppliers')

def initialize_bill_to_companies():
    cursor.execute("SELECT COUNT(*) FROM bill_to_companies")
    count = cursor.fetchone()[0]
    
    if count == 0:
        bill_to_data = [
            ("Zerone Technologies Pvt Ltd", "R/O No-350,Rajpur Khurd Extension,Near Zara Farm House Chhatapur, New Delhi 110074", "07AAACZ5805D1ZI", "", ""),
            ("QuantumInnovation Pvt Ltd", "123 Innovation Road, Pondicherry", "29ABCDE1234F2Z5", "", "")
        ]
        
        for company in bill_to_data:
            cursor.execute("""
                INSERT INTO bill_to_companies (company_name, address, gst_number, contact_person, contact_number)
                VALUES (%s, %s, %s, %s, %s)
            """, company)
        conn.commit()
        db_manager.backup_table('bill_to_companies')

def initialize_ship_to_addresses():
    cursor.execute("SELECT COUNT(*) FROM ship_to_addresses")
    count = cursor.fetchone()[0]
    
    if count == 0:
        ship_to_data = [
            ("Zerone Technologies Pvt Ltd", "Tata Projects Limited_ANANT UNIV.ANANT NATIONAL UNIVERSITYSANSKARDHAM CAMPUS, BOPAL-GUMASANAND ROAD AHMEDABAD Gujarat - 382115", "24AAACT4119L1Z", "", "91 89745 15576"),
            ("Mr. Dinesh", "45 Delivery Lane, Bangalore", "33XYZ7890K2", "Mr. Dinesh", "+91-9123456789")
        ]
        
        for address in ship_to_data:
            cursor.execute("""
                INSERT INTO ship_to_addresses (name, address, gst_number, contact_person, contact_number)
                VALUES (%s, %s, %s, %s, %s)
            """, address)
        conn.commit()
        db_manager.backup_table('ship_to_addresses')

def initialize_locations():
    cursor.execute("SELECT COUNT(*) FROM locations")
    count = cursor.fetchone()[0]
    
    if count == 0:
        locations_data = [("HR", "Haryana"), ("DL", "Delhi"), ("PN", "Pune")]
        
        for location_code, location_name in locations_data:
            cursor.execute("""
                INSERT INTO locations (location_code, location_name)
                VALUES (%s, %s)
            """, (location_code, location_name))
        conn.commit()
        db_manager.backup_table('locations')

def initialize_po_counters():
    cursor.execute("SELECT location_code FROM locations")
    locations = cursor.fetchall()
    
    for (location_code,) in locations:
        cursor.execute("SELECT COUNT(*) FROM po_counters WHERE location_code = %s", (location_code,))
        exists = cursor.fetchone()[0]
        
        if exists == 0:
            cursor.execute("""
                INSERT INTO po_counters (location_code, last_serial_number)
                VALUES (%s, %s)
            """, (location_code, 0))
    
    conn.commit()
    db_manager.backup_table('po_counters')

# Helper functions
def get_all_suppliers():
    cursor.execute("SELECT id, name, address, gst_number, contact_person, contact_number FROM suppliers ORDER BY name")
    return cursor.fetchall()

def get_all_bill_to_companies():
    cursor.execute("SELECT id, company_name, address, gst_number, contact_person, contact_number FROM bill_to_companies ORDER BY company_name")
    return cursor.fetchall()

def get_all_ship_to_addresses():
    cursor.execute("SELECT id, name, address, gst_number, contact_person, contact_number FROM ship_to_addresses ORDER BY name")
    return cursor.fetchall()

def get_all_locations():
    cursor.execute("SELECT location_code, location_name FROM locations ORDER BY location_name")
    return cursor.fetchall()

def get_current_financial_year():
    """Get current Indian Financial Year in 2K25-2K26 format"""
    today = datetime.datetime.now()
    
    if today.month >= 4:
        fy_start = today.year
        fy_end = today.year + 1
    else:
        fy_start = today.year - 1
        fy_end = today.year
    
    return f"2K{str(fy_start)[-2:]}-2K{str(fy_end)[-2:]}"

def generate_po_number(location_code):
    """Generate next PO number for given location"""
    fy_year = get_current_financial_year()
    
    cursor.execute("SELECT last_serial_number FROM po_counters WHERE location_code = %s", (location_code,))
    result = cursor.fetchone()
    
    if result:
        current_serial = result[0]
        next_serial = current_serial + 1
        
        cursor.execute("""
            UPDATE po_counters 
            SET last_serial_number = %s, updated_at = CURRENT_TIMESTAMP 
            WHERE location_code = %s
        """, (next_serial, location_code))
        conn.commit()
        db_manager.backup_table('po_counters')
    else:
        next_serial = 1
        cursor.execute("""
            INSERT INTO po_counters (location_code, last_serial_number)
            VALUES (%s, %s)
        """, (location_code, next_serial))
        conn.commit()
        db_manager.backup_table('po_counters')
    
    po_number = f"ZTPL-{location_code}/{fy_year}-{next_serial:03d}"
    return po_number

def clean_numeric(value):
    """Clean numeric values from strings with commas, spaces, etc."""
    if pd.isna(value) or value == '':
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(',', '').replace(' ', '')
    match = re.search(r'[\d.]+', cleaned)
    if match:
        return float(match.group())
    return 0

def generate_po_pdf(po_number, project_name, supplier_name, bill_to_name, ship_to_name, items, location_name):
    """Generate PDF for Purchase Order"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Header
    elements.append(Paragraph("Zerone Technologies Pvt Ltd", styles['Heading1']))
    elements.append(Paragraph(f"Purchase Order: {po_number}", styles['Title']))
    elements.append(Paragraph(f"Date: {datetime.datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Paragraph(f"Project: {project_name}", styles['Normal']))
    elements.append(Paragraph(f"Location: {location_name}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Get company details
    supplier = cursor.execute("SELECT * FROM suppliers WHERE name = %s", (supplier_name,)).fetchone()
    bill_to = cursor.execute("SELECT * FROM bill_to_companies WHERE company_name = %s", (bill_to_name,)).fetchone()
    ship_to = cursor.execute("SELECT * FROM ship_to_addresses WHERE name = %s", (ship_to_name,)).fetchone()

    # Company details sections
    if supplier:
        elements.append(Paragraph("Supplier:", styles['Heading2']))
        elements.append(Paragraph(f"{supplier[1]}<br/>{supplier[2]}<br/>GST: {supplier[3]}<br/>Contact: {supplier[4]} ({supplier[5]})", styles['Normal']))
        elements.append(Spacer(1, 12))

    if bill_to:
        elements.append(Paragraph("Bill To:", styles['Heading2']))
        elements.append(Paragraph(f"{bill_to[1]}<br/>{bill_to[2]}<br/>GST: {bill_to[3]}", styles['Normal']))
        elements.append(Spacer(1, 12))

    if ship_to:
        elements.append(Paragraph("Ship To:", styles['Heading2']))
        elements.append(Paragraph(f"{ship_to[1]}<br/>{ship_to[2]}<br/>GST: {ship_to[3]}<br/>Contact: {ship_to[5]}", styles['Normal']))
        elements.append(Spacer(1, 12))

    # Items table
    data = [['BOQ Ref', 'Description', 'Unit', 'Quantity', 'Rate', 'Amount']]
    total = 0
    for _, row in items.iterrows():
        if row.get('Quantity', 0) > 0:
            amount = row.get('Quantity', 0) * row.get('Unit Price', row.get('rate', 0))
            total += amount
            data.append([
                row.get('boq_ref', ''),
                row.get('description', ''),
                row.get('unit', ''),
                str(row.get('Quantity', 0)),
                f"₹{row.get('Unit Price', row.get('rate', 0)):.2f}",
                f"₹{amount:.2f}"
            ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)

    # Total
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Total Amount: ₹{total:.2f}", styles['Heading2']))

    # Terms and Conditions
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Terms and Conditions:", styles['Heading2']))
    elements.append(Paragraph("1. Delivery within 30 days.<br/>2. Payment terms: 30 days net.<br/>3. All disputes subject to local jurisdiction.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    return buffer

# Main application
def main_app():
    # Header with user info and backup controls
    col1, col2, col3, col4, col5 = st.columns([3, 1, 1, 1, 1])
    with col1:
        st.title("📦 BOQ & Purchase Order Management System")
        st.caption(f"Welcome, {st.session_state.get('user_name', st.session_state['username'])} ({st.session_state['role']})")
    
    with col2:
        if st.button("💾 Manual Backup"):
            with st.spinner("Creating backup..."):
                backup_now()
            st.success("✅ Backup completed!")
            st.rerun()
    
    with col3:
        if st.button("📊 Backup Status"):
            status = get_backup_status()
            st.info(f"Desktop: {status['desktop_files']} files\nServer: {status['server_files']} files\nStatus: {status['server_status']}")
    
    with col4:
        if st.button("🔗 Test Server"):
            if test_server_connection():
                st.success("✅ Server OK")
            else:
                st.error("❌ Server Error")
    
    with col5:
        if st.button("🚪 Logout"):
            logout()

    # Create and initialize all tables
    create_projects_table()
    create_boq_items_table()
    create_suppliers_table()
    initialize_suppliers()
    create_bill_to_table()
    initialize_bill_to_companies()
    create_ship_to_table()
    initialize_ship_to_addresses()
    create_locations_table()
    initialize_locations()
    create_po_counters_table()
    initialize_po_counters()

    # Main navigation tabs - Restrict Company Management for Staff
    main_tabs = ["📤 BOQ Management", "📋 View BOQ Items", "📄 Generate Purchase Order"]
    
    # Only Admin can access Company Management and User Management
    if st.session_state['role'] == 'admin':
        main_tabs.extend(["👥 Manage Companies", "👤 User Management"])
    
    selected_tab = st.selectbox("Select Function", main_tabs, key="main_navigation")

    # TAB 1: BOQ Management
    if selected_tab == "📤 BOQ Management":
        st.subheader("📤 Upload BOQ Excel/CSV & Create Project")
        
        project_name = st.text_input("Enter New Project Name")
        uploaded_file = st.file_uploader("Upload BOQ File", type=["xlsx", "xlsm", "csv"])
        
        if project_name and uploaded_file and st.button("🚀 Upload & Save BOQ"):
            try:
                # Read file based on extension
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    excel_sheets = pd.read_excel(uploaded_file, sheet_name=None)
                    
                    target_sheet = None
                    for sheet_name, sheet_df in excel_sheets.items():
                        if not sheet_df.empty and 'BOQ Ref' in sheet_df.columns:
                            target_sheet = sheet_name
                            df = sheet_df
                            break
                    
                    if target_sheet is None:
                        for sheet_name, sheet_df in excel_sheets.items():
                            if 'PROJECT' in sheet_name.upper() or 'BOQ' in sheet_name.upper():
                                target_sheet = sheet_name
                                df = sheet_df
                                break
                    
                    if target_sheet is None:
                        target_sheet = list(excel_sheets.keys())[0]
                        df = excel_sheets[target_sheet]
                    
                    st.info(f"📊 Reading data from sheet: *{target_sheet}*")
                
                # Column mapping for different formats
                column_mapping = {
                    'BOQ Ref': 'boq_ref', 'Description': 'description', 'Make': 'make', 'Model': 'model',
                    'Unit': 'unit', 'BOQ Qty.': 'boq_qty', 'Rate': 'rate', 'Amount': 'amount',
                    'BOQ Qty': 'boq_qty', 'boq ref': 'boq_ref', 'boq_ref': 'boq_ref',
                    'description': 'description', 'make': 'make', 'model': 'model',
                    'unit': 'unit', 'boq qty': 'boq_qty', 'boq_qty': 'boq_qty',
                    'rate': 'rate', 'amount': 'amount'
                }
                
                # Apply column mapping
                matched_columns = {}
                for actual_col in df.columns:
                    actual_col_str = str(actual_col).strip()
                    if actual_col_str in column_mapping:
                        matched_columns[actual_col] = column_mapping[actual_col_str]
                
                df = df.rename(columns=matched_columns)
                
                # Check required columns
                minimum_required = ['boq_ref', 'description', 'unit', 'boq_qty', 'rate']
                missing_cols = [col for col in minimum_required if col not in df.columns]
                
                if missing_cols:
                    st.error(f"❌ Missing required columns: {missing_cols}")
                else:
                    # Fill missing columns with defaults
                    for col in ['make', 'model']:
                        if col not in df.columns:
                            df[col] = 'N/A'
                    
                    # Create delivery quantity columns
                    for i in range(1, 11):
                        col_name = f'delivered_qty_{i}'
                        if col_name not in df.columns:
                            df[col_name] = 0
                    
                    # Clean and convert numeric columns
                    df['boq_qty'] = df['boq_qty'].apply(clean_numeric)
                    df['rate'] = df['rate'].apply(clean_numeric)
                    
                    if 'amount' not in df.columns:
                        df['amount'] = df['boq_qty'] * df['rate']
                    else:
                        df['amount'] = df['amount'].apply(clean_numeric)
                    
                    # Clean delivery quantities
                    for i in range(1, 11):
                        col_name = f'delivered_qty_{i}'
                        df[col_name] = df[col_name].apply(clean_numeric)
                    
                    # Calculate totals
                    df['total_delivery_qty'] = df[[f'delivered_qty_{i}' for i in range(1, 11)]].sum(axis=1)
                    df['balance_to_deliver'] = df['boq_qty'] - df['total_delivery_qty']
                    
                    df = df.fillna(0)
                    
                    # Insert project
                    cursor.execute("INSERT INTO projects (name, created_by) VALUES (%s, %s) RETURNING id", 
                                   (project_name, st.session_state['user_id']))
                    project_id = cursor.fetchone()[0]
                    
                    # Insert BOQ items
                    success_count = 0
                    error_count = 0
                    
                    for idx, row in df.iterrows():
                        try:
                            cursor.execute("""
                                INSERT INTO boq_items (
                                    project_id, boq_ref, description, make, model, unit, boq_qty, rate, amount,
                                    delivered_qty_1, delivered_qty_2, delivered_qty_3, delivered_qty_4, delivered_qty_5,
                                    delivered_qty_6, delivered_qty_7, delivered_qty_8, delivered_qty_9, delivered_qty_10,
                                    total_delivery_qty, balance_to_deliver
                                ) VALUES (
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s
                                )
                            """, (
                                project_id, str(row['boq_ref']), str(row['description']), str(row['make']), 
                                str(row['model']), str(row['unit']), float(row['boq_qty']), float(row['rate']), 
                                float(row['amount']), float(row['delivered_qty_1']), float(row['delivered_qty_2']), 
                                float(row['delivered_qty_3']), float(row['delivered_qty_4']), float(row['delivered_qty_5']), 
                                float(row['delivered_qty_6']), float(row['delivered_qty_7']), float(row['delivered_qty_8']), 
                                float(row['delivered_qty_9']), float(row['delivered_qty_10']), float(row['total_delivery_qty']), 
                                float(row['balance_to_deliver'])
                            ))
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            st.write(f"Error in row {idx}: {str(e)}")
                    
                    conn.commit()
                    
                    # Backup after BOQ upload
                    db_manager.backup_table('projects')
                    db_manager.backup_table('boq_items')
                    
                    st.success(f"✅ BOQ uploaded successfully! {success_count} items inserted, {error_count} errors.")
                    
                    # Show preview of processed data
                    st.subheader("📋 Preview of Processed Data")
                    st.dataframe(df.head(10))
                    
            except Exception as e:
                st.error(f"❌ Error while uploading BOQ: {str(e)}")
                st.write("Please check your file format and try again.")

        # Project deletion (Admin only)
        if st.session_state['role'] == 'admin':
            with st.expander("🗑️ Delete Project (Admin Only)"):
                cursor.execute("SELECT id, name FROM projects ORDER BY id DESC")
                projects = cursor.fetchall()
                if projects:
                    project_options = {name: pid for pid, name in projects}
                    project_to_delete = st.selectbox("Select Project to Delete", list(project_options.keys()))
                    
                    if st.button("🗑️ Delete Project", type="secondary"):
                        if st.button("⚠️ Confirm Delete", type="primary"):
                            project_id = project_options[project_to_delete]
                            cursor.execute("DELETE FROM projects WHERE id = %s", (project_id,))
                            conn.commit()
                            
                            db_manager.backup_table('projects')
                            db_manager.backup_table('boq_items')
                            
                            st.success(f"✅ Project '{project_to_delete}' deleted successfully!")
                            st.rerun()

    # TAB 2: View BOQ Items
    elif selected_tab == "📋 View BOQ Items":
        st.subheader("📋 View BOQ Items for Existing Project")
        cursor.execute("SELECT id, name FROM projects ORDER BY id DESC")
        projects = cursor.fetchall()

        if projects:
            project_options = {name: pid for pid, name in projects}
            selected_project = st.selectbox("Select a Project to View BOQ", list(project_options.keys()))

            if selected_project:
                project_id = project_options[selected_project]
                
                col1, col2 = st.columns([3, 1])
                with col2:
                    if st.session_state['role'] == 'admin' and st.button("🗑 Delete This Project"):
                        cursor.execute("DELETE FROM projects WHERE id = %s", (project_id,))
                        conn.commit()
                        
                        db_manager.backup_table('projects')
                        db_manager.backup_table('boq_items')
                        
                        st.success("✅ Project and its BOQ items deleted.")
                        st.rerun()
                
                cursor.execute("SELECT * FROM boq_items WHERE project_id = %s", (project_id,))
                records = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]

                boq_df = pd.DataFrame(records, columns=columns)

                if not boq_df.empty:
                    st.subheader("🔍 Search in BOQ Table")
                    search_term = st.text_input("Search by Description, Make, or Model")
                    if search_term:
                        mask = boq_df["description"].astype(str).str.contains(search_term, case=False) | \
                               boq_df["make"].astype(str).str.contains(search_term, case=False) | \
                               boq_df["model"].astype(str).str.contains(search_term, case=False)
                        filtered_df = boq_df[mask]
                        st.dataframe(filtered_df, use_container_width=True)
                    else:
                        st.dataframe(boq_df, use_container_width=True)
                else:
                    st.warning("⚠ No BOQ items found for this project.")
        else:
            st.info("ℹ No projects found. Upload a project first.")

    # TAB 3: Generate Purchase Order
    elif selected_tab == "📄 Generate Purchase Order":
        st.subheader("📄 Generate Purchase Order")
        
        cursor.execute("SELECT id, name FROM projects ORDER BY id DESC")
        projects = cursor.fetchall()
        
        if projects:
            project_options = {name: pid for pid, name in projects}
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.header("📌 PO Configuration")
                po_project = st.selectbox("Select Project for PO", list(project_options.keys()), key="po_project")
                po_project_id = project_options[po_project]
                
                st.subheader("📍 Location & PO Details")
                locations = get_all_locations()
                location_options = {f"{loc[1]} ({loc[0]})": loc[0] for loc in locations}
                
                selected_location_display = st.selectbox("🔽 Select Location for PO", list(location_options.keys()))
                selected_location_code = location_options[selected_location_display]
                
                if st.button("🔄 Generate New PO Number"):
                    auto_po_number = generate_po_number(selected_location_code)
                    st.session_state['generated_po_number'] = auto_po_number
                    st.success(f"✅ Generated PO Number: *{auto_po_number}*")
                
                if 'generated_po_number' in st.session_state:
                    current_po = st.session_state['generated_po_number']
                else:
                    preview_fy = get_current_financial_year()
                    cursor.execute("SELECT last_serial_number FROM po_counters WHERE location_code = %s", (selected_location_code,))
                    result = cursor.fetchone()
                    next_serial = (result[0] + 1) if result else 1
                    current_po = f"ZTPL-{selected_location_code}/{preview_fy}-{next_serial:03d}"
                    st.info(f"📋 Next PO Number will be: *{current_po}*")
                
                po_number = st.text_input("PO Number", value=current_po)
                po_date = st.date_input("PO Date", datetime.date.today())
                
                auto_reference = f"Ref#{selected_location_code}-PROJ-{po_project_id:02d}"
                po_reference = st.text_input("Reference", value=auto_reference)
                
                current_fy = get_current_financial_year()
                st.info(f"📅 Current Financial Year: *{current_fy}* (Indian FY: April-March)")
                
                selected_slot = st.selectbox("Select Delivery Slot", [f"delivered_qty_{i}" for i in range(1, 11)])
                
            with col2:
                st.header("📌 Company Details")
                
                # Supplier Details with Auto-fill
                st.subheader("🏢 Supplier Details")
                suppliers = get_all_suppliers()
                supplier_names = ["-- Select Supplier --"] + [supplier[1] for supplier in suppliers]
                
                selected_supplier_name = st.selectbox("🔽 Select Supplier", supplier_names)
                
                supplier_name = ""
                supplier_address = ""
                supplier_gst = ""
                supplier_contact = ""
                supplier_person = ""
                
                if selected_supplier_name != "-- Select Supplier --":
                    selected_supplier = next((s for s in suppliers if s[1] == selected_supplier_name), None)
                    if selected_supplier:
                        supplier_name = selected_supplier[1]
                        supplier_address = selected_supplier[2] or ""
                        supplier_gst = selected_supplier[3] or ""
                        supplier_person = selected_supplier[4] or ""
                        supplier_contact = selected_supplier[5] or ""
                        st.success(f"✅ Auto-filled details for: *{supplier_name}*")
                
                supplier_name = st.text_input("Supplier Name", value=supplier_name, key="supplier_name_input")
                supplier_address = st.text_area("Supplier Address", value=supplier_address, key="supplier_address_input")
                supplier_gst = st.text_input("Supplier GST No.", value=supplier_gst, key="supplier_gst_input")
                supplier_contact = st.text_input("Supplier Contact No.", value=supplier_contact, key="supplier_contact_input")
                supplier_person = st.text_input("Contact Person", value=supplier_person, key="supplier_person_input")
                
                # Bill To Details with Auto-fill
                st.subheader("📋 Bill To Details")
                bill_to_companies = get_all_bill_to_companies()
                bill_to_names = ["-- Select Bill To Company --"] + [company[1] for company in bill_to_companies]
                
                selected_bill_to_name = st.selectbox("🔽 Select Bill To Company", bill_to_names)
                
                bill_to_company = ""
                bill_to_address = ""
                bill_to_gst = ""
                
                if selected_bill_to_name != "-- Select Bill To Company --":
                    selected_bill_to = next((c for c in bill_to_companies if c[1] == selected_bill_to_name), None)
                    if selected_bill_to:
                        bill_to_company = selected_bill_to[1]
                        bill_to_address = selected_bill_to[2] or ""
                        bill_to_gst = selected_bill_to[3] or ""
                        st.success(f"✅ Auto-filled Bill To details for: *{bill_to_company}*")
                
                bill_to_company = st.text_input("Bill To: Company Name", value=bill_to_company, key="bill_to_company_input")
                bill_to_address = st.text_area("Bill To: Address", value=bill_to_address, key="bill_to_address_input")
                bill_to_gst = st.text_input("Bill To: GST No.", value=bill_to_gst, key="bill_to_gst_input")
                
                # Ship To Details with Auto-fill
                st.subheader("🚚 Ship To Details")
                ship_to_addresses = get_all_ship_to_addresses()
                ship_to_names = ["-- Select Ship To Address --"] + [address[1] for address in ship_to_addresses]
                
                selected_ship_to_name = st.selectbox("🔽 Select Ship To Address", ship_to_names)
                
                ship_to_name = ""
                ship_to_address = ""
                ship_to_gst = ""
                ship_to_contact = ""
                
                if selected_ship_to_name != "-- Select Ship To Address --":
                    selected_ship_to = next((a for a in ship_to_addresses if a[1] == selected_ship_to_name), None)
                    if selected_ship_to:
                        ship_to_name = selected_ship_to[1]
                        ship_to_address = selected_ship_to[2] or ""
                        ship_to_gst = selected_ship_to[3] or ""
                        ship_to_contact = selected_ship_to[5] or ""
                        st.success(f"✅ Auto-filled Ship To details for: *{ship_to_name}*")
                
                ship_to_name = st.text_input("Ship To: Name", value=ship_to_name, key="ship_to_name_input")
                ship_to_address = st.text_area("Ship To: Address", value=ship_to_address, key="ship_to_address_input")
                ship_to_gst = st.text_input("Ship To: GST No.", value=ship_to_gst, key="ship_to_gst_input")
                ship_to_contact = st.text_input("Ship To: Contact No.", value=ship_to_contact, key="ship_to_contact_input")
            
            # Logo and signature uploads
            st.subheader("📎 Upload Files")
            col3, col4 = st.columns(2)
            with col3:
                logo_file = st.file_uploader("Upload Company Logo", type=["png", "jpg", "jpeg"])
            with col4:
                sign_file = st.file_uploader("Upload Prepared By Signature", type=["png", "jpg", "jpeg"])
            
            # Get BOQ items for selected project
            cursor.execute("SELECT boq_ref, description, make, model, unit, CAST(rate as FLOAT) as rate, CAST(balance_to_deliver as FLOAT) as balance_to_deliver FROM boq_items WHERE project_id = %s", (po_project_id,))
            po_items = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            
            if po_items:
                po_df = pd.DataFrame(po_items, columns=columns)
                po_df["Quantity"] = 0.0
                po_df["Unit Price"] = po_df["rate"].astype(float)
                po_df["Delivery Slot"] = selected_slot
                
                st.subheader("📝 Edit Purchase Order Items")
                
                # Configure data editor permissions based on role
                if st.session_state['role'] == 'admin':
                    column_config = None  # Admin can edit all columns
                else:
                    column_config = {
                        "rate": st.column_config.NumberColumn(disabled=True),
                        "Unit Price": st.column_config.NumberColumn(disabled=True)
                    }
                
                updated_df = st.data_editor(po_df, use_container_width=True, num_rows="dynamic", key="po_editor", column_config=column_config)
                
                updated_df["Quantity"] = pd.to_numeric(updated_df["Quantity"], errors='coerce').fillna(0.0)
                updated_df["Unit Price"] = pd.to_numeric(updated_df["Unit Price"], errors='coerce').fillna(0.0)
                updated_df["Total"] = updated_df["Quantity"] * updated_df["Unit Price"]
                
                # Calculate totals
                subtotal = updated_df["Total"].sum()
                gst_percent = st.number_input("Enter GST %", min_value=0.0, value=18.0)
                gst_amount = (subtotal * gst_percent) / 100
                grand_total = subtotal + gst_amount
                
                try:
                    grand_total_words = f"{num2words(int(grand_total), lang='en_IN').title()} Rupees Only"
                except:
                    grand_total_words = f"{num2words(int(grand_total)).title()} Rupees Only"
                
                # Display totals
                st.markdown(f"**Subtotal:** ₹ {subtotal:,.2f}")
                st.markdown(f"**GST ({gst_percent}%):** ₹ {gst_amount:,.2f}")
                st.markdown(f"**Grand Total:** ₹ {grand_total:,.2f}")
                st.markdown(f"**In Words:** {grand_total_words}")
                
                # Terms and Conditions
                st.subheader("📄 Terms & Conditions")
                terms = st.text_area("Enter Terms & Conditions", height=150, value="""1. Payment due within 30 days.
2. Items are covered under manufacturer warranty.
3. Delivery subject to stock availability.
4. All disputes subject to Chennai jurisdiction.""")
                
                # Generate PO buttons
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("💾 Generate Purchase Order Excel", use_container_width=True):
                        if not all([supplier_name.strip(), supplier_address.strip(), bill_to_company.strip(), 
                                   bill_to_address.strip(), ship_to_name.strip(), ship_to_address.strip()]):
                            st.error("❌ Please fill in all required company details!")
                        else:
                            # Validation for BOQ items
                            validation_failed = False
                            error_rows = []
                            
                            for _, row in updated_df.iterrows():
                                boq_ref = row["boq_ref"]
                                quantity = float(row["Quantity"])
                                unit_price = float(row["Unit Price"])
                                balance = float(row["balance_to_deliver"]) if hasattr(row["balance_to_deliver"], 'quantize') else float(row["balance_to_deliver"])
                                rate = float(row["rate"]) if hasattr(row["rate"], 'quantize') else float(row["rate"])
                                
                                if quantity > 0:
                                    if quantity > balance:
                                        validation_failed = True
                                        error_rows.append(f"{boq_ref} (Balance: {balance}, Tried: {quantity})")
                                    elif unit_price > rate * 1.10:
                                        validation_failed = True
                                        error_rows.append(f"{boq_ref} (Allowed Rate: ₹{rate * 1.10:.2f}, Entered: ₹{unit_price:.2f})")
                            
                            if validation_failed:
                                st.error("❌ Cannot proceed. Issues in the following items:\n" + "\n".join(error_rows))
                            else:
                                # Update database with delivered quantities
                                for _, row in updated_df.iterrows():
                                    if float(row["Quantity"]) > 0:
                                        boq_ref = row["boq_ref"]
                                        quantity = float(row["Quantity"])
                                        
                                        cursor.execute(f"""
                                            SELECT id, boq_qty, total_delivery_qty, balance_to_deliver,
                                                delivered_qty_1, delivered_qty_2, delivered_qty_3, delivered_qty_4, delivered_qty_5,
                                                delivered_qty_6, delivered_qty_7, delivered_qty_8, delivered_qty_9, delivered_qty_10
                                            FROM boq_items
                                            WHERE project_id = %s AND boq_ref = %s
                                        """, (po_project_id, boq_ref))
                                        result = cursor.fetchone()
                                        
                                        if result:
                                            item_id = result[0]
                                            boq_qty = result[1]
                                            delivered_list = list(result[4:])
                                            slot_index = int(selected_slot.split("_")[-1]) - 1
                                            
                                            delivered_list[slot_index] += Decimal(str(quantity))
                                            total_delivered = sum(delivered_list)
                                            new_balance = boq_qty - total_delivered
                                            
                                            cursor.execute(f"""
                                                UPDATE boq_items SET
                                                    delivered_qty_1 = %s, delivered_qty_2 = %s, delivered_qty_3 = %s,
                                                    delivered_qty_4 = %s, delivered_qty_5 = %s, delivered_qty_6 = %s,
                                                    delivered_qty_7 = %s, delivered_qty_8 = %s, delivered_qty_9 = %s,
                                                    delivered_qty_10 = %s, total_delivery_qty = %s, balance_to_deliver = %s
                                                WHERE id = %s
                                            """, (*delivered_list, total_delivered, new_balance, item_id))
                                
                                conn.commit()
                                db_manager.backup_table('boq_items')
                                
                                # Generate Excel PO with full implementation
                                wb = Workbook()
                                ws = wb.active
                                ws.title = "Purchase Order"
                                row = 1
                                
                                # Define colors and styles
                                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                                title_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
                                total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                                
                                # Define border styles
                                thin_border = Border(
                                    left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin')
                                )
                                
                                # Set column widths
                                optimized_widths = {'A': 5, 'B': 35, 'C': 10, 'D': 12, 'E': 5, 'F': 6, 'G': 8, 'H': 10}
                                for col_letter, width in optimized_widths.items():
                                    ws.column_dimensions[col_letter].width = width
                                
                                # Logo (if uploaded)
                                if logo_file:
                                    try:
                                        img = Image.open(logo_file)
                                        img.thumbnail((60, 60))
                                        img_io = BytesIO()
                                        img.save(img_io, format="PNG")
                                        img_io.seek(0)
                                        ws.add_image(XLImage(img_io), "A1")
                                        row += 4
                                    except Exception as e:
                                        st.warning(f"Could not add logo: {str(e)}")
                                
                                # Company header
                                title_cell = ws.cell(row=row, column=1)
                                title_cell.value = "PURCHASE ORDER"
                                title_cell.font = Font(bold=True, size=16, color="FFFFFF")
                                title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                                title_cell.alignment = Alignment(horizontal='center')
                                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                                row += 2
                                
                                # PO Details
                                ws.cell(row=row, column=1).value = "PO Number:"
                                ws.cell(row=row, column=1).font = Font(bold=True)
                                ws.cell(row=row, column=2).value = po_number
                                ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
                                
                                ws.cell(row=row, column=5).value = "Date:"
                                ws.cell(row=row, column=5).font = Font(bold=True)
                                ws.cell(row=row, column=6).value = po_date.strftime("%d/%m/%Y")
                                row += 2
                                
                                # Supplier Details
                                ws.cell(row=row, column=1).value = "Supplier:"
                                ws.cell(row=row, column=1).font = Font(bold=True)
                                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                                ws.cell(row=row, column=2).value = supplier_name
                                row += 1
                                
                                ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=4)
                                addr_cell = ws.cell(row=row, column=2)
                                addr_cell.value = supplier_address
                                addr_cell.alignment = Alignment(wrap_text=True, vertical='top')
                                row += 2
                                
                                # Bill To Details
                                ws.cell(row=row, column=1).value = "Bill To:"
                                ws.cell(row=row, column=1).font = Font(bold=True)
                                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                                ws.cell(row=row, column=2).value = bill_to_company
                                row += 1
                                
                                ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=4)
                                bill_addr_cell = ws.cell(row=row, column=2)
                                bill_addr_cell.value = bill_to_address
                                bill_addr_cell.alignment = Alignment(wrap_text=True, vertical='top')
                                row += 2
                                
                                # Ship To Details
                                ws.cell(row=row, column=1).value = "Ship To:"
                                ws.cell(row=row, column=1).font = Font(bold=True)
                                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                                ws.cell(row=row, column=2).value = ship_to_name
                                row += 1
                                
                                ws.merge_cells(start_row=row, start_column=2, end_row=row+1, end_column=4)
                                ship_addr_cell = ws.cell(row=row, column=2)
                                ship_addr_cell.value = ship_to_address
                                ship_addr_cell.alignment = Alignment(wrap_text=True, vertical='top')
                                row += 3
                                
                                # Table Headers
                                headers = ["S.No", "Description", "Make", "Model", "Unit", "Qty", "Rate", "Amount"]
                                for col_num, header in enumerate(headers, 1):
                                    cell = ws.cell(row=row, column=col_num)
                                    cell.value = header
                                    cell.font = Font(bold=True, size=10, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                                    cell.alignment = Alignment(horizontal='center', vertical='center')
                                    cell.border = thin_border
                                row += 1
                                
                                # Product data
                                filtered_items = updated_df[updated_df["Quantity"] > 0]
                                for idx, (_, item) in enumerate(filtered_items.iterrows(), 1):
                                    ws.row_dimensions[row].height = 35
                                    
                                    # Data cells
                                    data_cells = [
                                        (1, idx),
                                        (2, item["description"]),
                                        (3, item["make"]),
                                        (4, item["model"]),
                                        (5, item["unit"]),
                                        (6, item["Quantity"]),
                                        (7, f"₹{item['Unit Price']:.2f}"),
                                        (8, f"₹{item['Total']:.2f}")
                                    ]
                                    
                                    for col, value in data_cells:
                                        cell = ws.cell(row=row, column=col)
                                        cell.value = value
                                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                        cell.font = Font(size=9)
                                        cell.border = thin_border
                                        if col == 2:  # Description
                                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    
                                    row += 1
                                
                                # Totals section
                                row += 1
                                
                                # Subtotal
                                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                                total_cell = ws.cell(row=row, column=1)
                                total_cell.value = "Sub Total"
                                total_cell.font = Font(bold=True, size=11)
                                total_cell.fill = total_fill
                                total_cell.alignment = Alignment(horizontal='right')
                                
                                subtotal_cell = ws.cell(row=row, column=8)
                                subtotal_cell.value = f"₹{subtotal:,.2f}"
                                subtotal_cell.font = Font(bold=True, size=11)
                                subtotal_cell.fill = total_fill
                                subtotal_cell.alignment = Alignment(horizontal='center')
                                row += 1
                                
                                # GST
                                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                                gst_cell = ws.cell(row=row, column=1)
                                gst_cell.value = f"GST ({gst_percent}%)"
                                gst_cell.font = Font(bold=True, size=10)
                                gst_cell.fill = total_fill
                                gst_cell.alignment = Alignment(horizontal='right')
                                
                                gst_amt_cell = ws.cell(row=row, column=8)
                                gst_amt_cell.value = f"₹{gst_amount:,.2f}"
                                gst_amt_cell.font = Font(bold=True, size=10)
                                gst_amt_cell.fill = total_fill
                                gst_amt_cell.alignment = Alignment(horizontal='center')
                                row += 1
                                
                                # Grand Total
                                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                                grand_label_cell = ws.cell(row=row, column=1)
                                grand_label_cell.value = "TOTAL:"
                                grand_label_cell.font = Font(bold=True, size=12)
                                grand_label_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                                grand_label_cell.alignment = Alignment(horizontal='center')
                                
                                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
                                words_cell = ws.cell(row=row, column=3)
                                words_cell.value = grand_total_words
                                words_cell.font = Font(bold=True, size=9)
                                words_cell.alignment = Alignment(horizontal='center')
                                
                                grand_total_cell = ws.cell(row=row, column=8)
                                grand_total_cell.value = f"₹{grand_total:,.2f}"
                                grand_total_cell.font = Font(bold=True, size=12, color="FF0000")
                                grand_total_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                                grand_total_cell.alignment = Alignment(horizontal='center')
                                
                                # Terms section
                                row += 2
                                terms_header_cell = ws.cell(row=row, column=1)
                                terms_header_cell.value = "TERMS & CONDITIONS:"
                                terms_header_cell.font = Font(bold=True, size=10)
                                terms_header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                                row += 1
                                
                                # Add terms
                                terms_lines = terms.split('\n')
                                for term_line in terms_lines:
                                    if term_line.strip():
                                        terms_cell = ws.cell(row=row, column=1)
                                        terms_cell.value = term_line.strip()
                                        terms_cell.font = Font(size=8)
                                        terms_cell.alignment = Alignment(wrap_text=True, vertical='top')
                                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
                                        ws.row_dimensions[row].height = 15
                                        row += 1
                                
                                # Save workbook
                                output = BytesIO()
                                wb.save(output)
                                output.seek(0)
                                
                                st.success("✅ Purchase Order Excel generated successfully!")
                                st.download_button(
                                    "📥 Download Purchase Order Excel",
                                    data=output.getvalue(),
                                    file_name=f"Purchase_Order_{po_number}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                
                with col2:
                    if st.button("📄 Generate Purchase Order PDF", use_container_width=True):
                        if not all([supplier_name.strip(), supplier_address.strip(), bill_to_company.strip(), 
                                   bill_to_address.strip(), ship_to_name.strip(), ship_to_address.strip()]):
                            st.error("❌ Please fill in all required company details!")
                        else:
                            # Validation for BOQ items
                            validation_failed = False
                            error_rows = []
                            
                            for _, row in updated_df.iterrows():
                                boq_ref = row["boq_ref"]
                                quantity = float(row["Quantity"])
                                unit_price = float(row["Unit Price"])
                                balance = float(row["balance_to_deliver"]) if hasattr(row["balance_to_deliver"], 'quantize') else float(row["balance_to_deliver"])
                                rate = float(row["rate"]) if hasattr(row["rate"], 'quantize') else float(row["rate"])
                                
                                if quantity > 0:
                                    if quantity > balance:
                                        validation_failed = True
                                        error_rows.append(f"{boq_ref} (Balance: {balance}, Tried: {quantity})")
                                    elif unit_price > rate * 1.10:
                                        validation_failed = True
                                        error_rows.append(f"{boq_ref} (Allowed Rate: ₹{rate * 1.10:.2f}, Entered: ₹{unit_price:.2f})")
                            
                            if validation_failed:
                                st.error("❌ Cannot proceed. Issues in the following items:\n" + "\n".join(error_rows))
                            else:
                                # Update database with delivered quantities
                                for _, row in updated_df.iterrows():
                                    if float(row["Quantity"]) > 0:
                                        boq_ref = row["boq_ref"]
                                        quantity = float(row["Quantity"])
                                        
                                        cursor.execute(f"""
                                            SELECT id, boq_qty, total_delivery_qty, balance_to_deliver,
                                                delivered_qty_1, delivered_qty_2, delivered_qty_3, delivered_qty_4, delivered_qty_5,
                                                delivered_qty_6, delivered_qty_7, delivered_qty_8, delivered_qty_9, delivered_qty_10
                                            FROM boq_items
                                            WHERE project_id = %s AND boq_ref = %s
                                        """, (po_project_id, boq_ref))
                                        result = cursor.fetchone()
                                        
                                        if result:
                                            item_id = result[0]
                                            boq_qty = result[1]
                                            delivered_list = list(result[4:])
                                            slot_index = int(selected_slot.split("_")[-1]) - 1
                                            
                                            delivered_list[slot_index] += Decimal(str(quantity))
                                            total_delivered = sum(delivered_list)
                                            new_balance = boq_qty - total_delivered
                                            
                                            cursor.execute(f"""
                                                UPDATE boq_items SET
                                                    delivered_qty_1 = %s, delivered_qty_2 = %s, delivered_qty_3 = %s,
                                                    delivered_qty_4 = %s, delivered_qty_5 = %s, delivered_qty_6 = %s,
                                                    delivered_qty_7 = %s, delivered_qty_8 = %s, delivered_qty_9 = %s,
                                                    delivered_qty_10 = %s, total_delivery_qty = %s, balance_to_deliver = %s
                                                WHERE id = %s
                                            """, (*delivered_list, total_delivered, new_balance, item_id))
                                
                                conn.commit()
                                db_manager.backup_table('boq_items')
                                
                                # Generate PDF PO
                                pdf_buffer = generate_po_pdf(po_number, po_project, supplier_name, bill_to_company, 
                                                           ship_to_name, updated_df, selected_location_display)
                                
                                st.success("✅ Purchase Order PDF generated successfully!")
                                st.download_button(
                                    label="📥 Download Purchase Order PDF",
                                    data=pdf_buffer.getvalue(),
                                    file_name=f"Purchase_Order_{po_number}.pdf",
                                    mime="application/pdf"
                                )
            else:
                st.warning("⚠ No BOQ items found for this project.")
        else:
            st.warning("⚠ No projects found. Please create a project first in the BOQ Management tab.")

    # TAB 4: Manage Companies
    elif selected_tab == "👥 Manage Companies":
        st.subheader("👥 Company Management")
        
        subtab1, subtab2, subtab3, subtab4, subtab5 = st.tabs(["🏢 Suppliers", "📋 Bill To Companies", "🚚 Ship To Addresses", "📍 Locations", "💾 Backup Center"])
        
        # SUPPLIERS MANAGEMENT
        with subtab1:
            st.header("🏢 Supplier Management")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("➕ Add New Supplier")
                
                with st.form("add_supplier_form"):
                    new_supplier_name = st.text_input("Supplier Name*")
                    new_supplier_address = st.text_area("Supplier Address")
                    new_supplier_gst = st.text_input("GST Number")
                    new_supplier_person = st.text_input("Contact Person")
                    new_supplier_contact = st.text_input("Contact Number")
                    
                    submit_supplier = st.form_submit_button("💾 Add Supplier")
                    
                    if submit_supplier:
                        if new_supplier_name.strip():
                            try:
                                cursor.execute("""
                                    INSERT INTO suppliers (name, address, gst_number, contact_person, contact_number)
                                    VALUES (%s, %s, %s, %s, %s)
                                """, (
                                    new_supplier_name.strip(),
                                    new_supplier_address.strip(),
                                    new_supplier_gst.strip(),
                                    new_supplier_person.strip(),
                                    new_supplier_contact.strip()
                                ))
                                conn.commit()
                                db_manager.backup_table('suppliers')
                                st.success(f"✅ Supplier '{new_supplier_name}' added successfully!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error adding supplier: {str(e)}")
                        else:
                            st.error("❌ Supplier name is required!")
            
            with col2:
                st.subheader("📋 Existing Suppliers")
                
                suppliers = get_all_suppliers()
                
                if suppliers:
                    suppliers_df = pd.DataFrame(suppliers, columns=[
                        'ID', 'Name', 'Address', 'GST Number', 'Contact Person', 'Contact Number'
                    ])
                    
                    search_supplier = st.text_input("🔍 Search Suppliers", key="search_supplier")
                    
                    if search_supplier:
                        mask = suppliers_df['Name'].str.contains(search_supplier, case=False, na=False) | \
                               suppliers_df['Address'].str.contains(search_supplier, case=False, na=False)
                        filtered_suppliers = suppliers_df[mask]
                    else:
                        filtered_suppliers = suppliers_df
                    
                    for idx, supplier in filtered_suppliers.iterrows():
                        with st.expander(f"🏢 {supplier['Name']}", expanded=False):
                            st.write(f"**Address:** {supplier['Address']}")
                            st.write(f"**GST:** {supplier['GST Number']}")
                            st.write(f"**Contact Person:** {supplier['Contact Person']}")
                            st.write(f"**Contact:** {supplier['Contact Number']}")
                            
                            if st.session_state['role'] == 'admin' and st.button(f"🗑 Delete", key=f"delete_supplier_{supplier['ID']}"):
                                try:
                                    cursor.execute("DELETE FROM suppliers WHERE id = %s", (supplier['ID'],))
                                    conn.commit()
                                    db_manager.backup_table('suppliers')
                                    st.success(f"✅ Supplier '{supplier['Name']}' deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error deleting supplier: {str(e)}")
                else:
                    st.info("ℹ No suppliers found. Add some suppliers to get started!")

        # BILL TO COMPANIES MANAGEMENT
        with subtab2:
            st.header("📋 Bill To Companies Management")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("➕ Add New Bill To Company")
                
                with st.form("add_bill_to_form"):
                    new_company_name = st.text_input("Company Name*")
                    new_company_address = st.text_area("Company Address")
                    new_company_gst = st.text_input("GST Number")
                    new_company_person = st.text_input("Contact Person")
                    new_company_contact = st.text_input("Contact Number")
                    
                    submit_bill_to = st.form_submit_button("💾 Add Bill To Company")
                    
                    if submit_bill_to:
                        if new_company_name.strip():
                            try:
                                cursor.execute("""
                                    INSERT INTO bill_to_companies (company_name, address, gst_number, contact_person, contact_number)
                                    VALUES (%s, %s, %s, %s, %s)
                                """, (
                                    new_company_name.strip(),
                                    new_company_address.strip(),
                                    new_company_gst.strip(),
                                    new_company_person.strip(),
                                    new_company_contact.strip()
                                ))
                                conn.commit()
                                db_manager.backup_table('bill_to_companies')
                                st.success(f"✅ Bill To company '{new_company_name}' added successfully!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error adding Bill To company: {str(e)}")
                        else:
                            st.error("❌ Company name is required!")
            
            with col2:
                st.subheader("📋 Existing Bill To Companies")
                
                bill_to_companies = get_all_bill_to_companies()
                
                if bill_to_companies:
                    for company in bill_to_companies:
                        with st.expander(f"🏢 {company[1]}", expanded=False):
                            st.write(f"**Address:** {company[2]}")
                            st.write(f"**GST:** {company[3]}")
                            st.write(f"**Contact Person:** {company[4]}")
                            st.write(f"**Contact:** {company[5]}")
                            
                            if st.session_state['role'] == 'admin' and st.button(f"🗑 Delete", key=f"delete_bill_to_{company[0]}"):
                                try:
                                    cursor.execute("DELETE FROM bill_to_companies WHERE id = %s", (company[0],))
                                    conn.commit()
                                    db_manager.backup_table('bill_to_companies')
                                    st.success(f"✅ Bill To company '{company[1]}' deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error deleting Bill To company: {str(e)}")

        # SHIP TO ADDRESSES MANAGEMENT
        with subtab3:
            st.header("🚚 Ship To Addresses Management")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("➕ Add New Ship To Address")
                
                with st.form("add_ship_to_form"):
                    new_ship_name = st.text_input("Ship To Name*")
                    new_ship_address = st.text_area("Ship To Address")
                    new_ship_gst = st.text_input("GST Number")
                    new_ship_person = st.text_input("Contact Person")
                    new_ship_contact = st.text_input("Contact Number")
                    
                    submit_ship_to = st.form_submit_button("💾 Add Ship To Address")
                    
                    if submit_ship_to:
                        if new_ship_name.strip():
                            try:
                                cursor.execute("""
                                    INSERT INTO ship_to_addresses (name, address, gst_number, contact_person, contact_number)
                                    VALUES (%s, %s, %s, %s, %s)
                                """, (
                                    new_ship_name.strip(),
                                    new_ship_address.strip(),
                                    new_ship_gst.strip(),
                                    new_ship_person.strip(),
                                    new_ship_contact.strip()
                                ))
                                conn.commit()
                                db_manager.backup_table('ship_to_addresses')
                                st.success(f"✅ Ship To address '{new_ship_name}' added successfully!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"❌ Error adding Ship To address: {str(e)}")
                        else:
                            st.error("❌ Ship To name is required!")
            
            with col2:
                st.subheader("📋 Existing Ship To Addresses")
                
                ship_to_addresses = get_all_ship_to_addresses()
                
                if ship_to_addresses:
                    for address in ship_to_addresses:
                        with st.expander(f"🚚 {address[1]}", expanded=False):
                            st.write(f"**Address:** {address[2]}")
                            st.write(f"**GST:** {address[3]}")
                            st.write(f"**Contact Person:** {address[4]}")
                            st.write(f"**Contact:** {address[5]}")
                            
                            if st.session_state['role'] == 'admin' and st.button(f"🗑 Delete", key=f"delete_ship_to_{address[0]}"):
                                try:
                                    cursor.execute("DELETE FROM ship_to_addresses WHERE id = %s", (address[0],))
                                    conn.commit()
                                    db_manager.backup_table('ship_to_addresses')
                                    st.success(f"✅ Ship To address '{address[1]}' deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error deleting Ship To address: {str(e)}")

        # LOCATIONS MANAGEMENT
        with subtab4:
            st.header("📍 Location Management")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("➕ Add New Location")
                
                with st.form("add_location_form"):
                    new_location_code = st.text_input("Location Code* (e.g., MH, KL, TN)", max_chars=10).upper()
                    new_location_name = st.text_input("Location Name* (e.g., Maharashtra, Kerala)")
                    
                    submit_location = st.form_submit_button("💾 Add Location")
                    
                    if submit_location:
                        if new_location_code.strip() and new_location_name.strip():
                            try:
                                cursor.execute("""
                                    INSERT INTO locations (location_code, location_name)
                                    VALUES (%s, %s)
                                """, (new_location_code.strip(), new_location_name.strip()))
                                
                                cursor.execute("""
                                    INSERT INTO po_counters (location_code, last_serial_number)
                                    VALUES (%s, %s)
                                """, (new_location_code.strip(), 0))
                                
                                conn.commit()
                                db_manager.backup_table('locations')
                                db_manager.backup_table('po_counters')
                                
                                st.success(f"✅ Location '{new_location_code} - {new_location_name}' added successfully!")
                                st.rerun()
                            except Exception as e:
                                if "duplicate key value" in str(e).lower():
                                    st.error(f"❌ Location code '{new_location_code}' already exists!")
                                else:
                                    st.error(f"❌ Error adding location: {str(e)}")
                        else:
                            st.error("❌ Both location code and name are required!")
            
            with col2:
                st.subheader("📋 Existing Locations")
                
                locations = get_all_locations()
                
                if locations:
                    location_counters = {}
                    for loc_code, loc_name in locations:
                        cursor.execute("SELECT last_serial_number FROM po_counters WHERE location_code = %s", (loc_code,))
                        result = cursor.fetchone()
                        location_counters[loc_code] = result[0] if result else 0
                    
                    current_fy = get_current_financial_year()
                    
                    for loc_code, loc_name in locations:
                        with st.expander(f"📍 {loc_name} ({loc_code})", expanded=False):
                            counter = location_counters[loc_code]
                            st.write(f"**Location Code:** {loc_code}")
                            st.write(f"**Location Name:** {loc_name}")
                            st.write(f"**Total POs Generated:** {counter}")
                            st.write(f"**Next PO Number:** ZTPL-{loc_code}/{current_fy}-{counter+1:03d}")
                            
                            if st.session_state['role'] == 'admin':
                                st.warning("⚠️ Deleting a location will affect PO number generation!")
                                if st.button(f"🗑️ Delete Location", key=f"delete_location_{loc_code}"):
                                    try:
                                        cursor.execute("DELETE FROM po_counters WHERE location_code = %s", (loc_code,))
                                        cursor.execute("DELETE FROM locations WHERE location_code = %s", (loc_code,))
                                        conn.commit()
                                        
                                        db_manager.backup_table('locations')
                                        db_manager.backup_table('po_counters')
                                        
                                        st.success(f"✅ Location '{loc_code} - {loc_name}' deleted!")
                                        st.rerun()
                                    except Exception as e:
                                        st.error(f"❌ Error deleting location: {str(e)}")
                else:
                    st.info("ℹ️ No locations found. Add some locations to get started!")

        # BACKUP CENTER
        with subtab5:
            st.header("💾 Backup Center")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("🎯 Manual Backup Operations")
                
                if st.button("📦 Backup All Tables", use_container_width=True):
                    with st.spinner("Creating complete backup..."):
                        backup_now()
                    st.success("✅ Complete backup finished!")
                
                st.subheader("📋 Individual Table Backups")
                
                backup_buttons = [
                    ("🏗 Projects", "projects"),
                    ("📋 BOQ Items", "boq_items"),
                    ("🏢 Suppliers", "suppliers"),
                    ("📄 Bill To Companies", "bill_to_companies"),
                    ("🚚 Ship To Addresses", "ship_to_addresses"),
                    ("📍 Locations", "locations"),
                    ("🔢 PO Counters", "po_counters")
                ]
                
                for label, table in backup_buttons:
                    if st.button(f"Backup {label}", key=f"backup_{table}"):
                        with st.spinner(f"Backing up {table}..."):
                            db_manager.backup_table(table)
                        st.success(f"✅ {label} backed up!")
            
            with col2:
                st.subheader("📊 Backup Status")
                
                status = get_backup_status()
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Desktop Files", status['desktop_files'])
                with col2:
                    st.metric("Server Files", status['server_files'])
                
                if "✅" in status['server_status']:
                    st.success(f"Server Status: {status['server_status']}")
                else:
                    st.error(f"Server Status: {status['server_status']}")
                
                st.info(f"Last Check: {status['last_backup']}")
                
                st.subheader("🔧 Connection Tests")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("Test Server Connection"):
                        if test_server_connection():
                            st.success("✅ Server connection successful!")
                        else:
                            st.error("❌ Server connection failed!")
                
                with col2:
                    if st.button("Test Database Connection"):
                        try:
                            test_conn = get_connection()
                            test_conn.close()
                            st.success("✅ Database connection successful!")
                        except Exception as e:
                            st.error(f"❌ Database connection failed: {e}")

    # TAB 5: User Management (Admin Only)
    elif selected_tab == "👤 User Management" and st.session_state['role'] == 'admin':
        st.subheader("👤 User Management (Admin Only)")
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.header("➕ Add New User")
            
            with st.form("add_user_form"):
                new_username = st.text_input("Username*")
                new_password = st.text_input("Password*", type="password")
                new_role = st.selectbox("Role*", ["admin", "staff"])
                new_name = st.text_input("Full Name")
                new_email = st.text_input("Email")
                new_contact = st.text_input("Contact Number")
                
                submit_user = st.form_submit_button("💾 Add User")
                
                if submit_user:
                    if new_username.strip() and new_password.strip():
                        try:
                            hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
                            
                            with auth_engine.connect() as auth_conn:
                                auth_conn.execute(text("""
                                    INSERT INTO users (username, password_hash, role, name, email, contact_number)
                                    VALUES (:username, :password_hash, :role, :name, :email, :contact_number)
                                """), {
                                    'username': new_username.strip(),
                                    'password_hash': hashed_password,
                                    'role': new_role,
                                    'name': new_name.strip(),
                                    'email': new_email.strip(),
                                    'contact_number': new_contact.strip()
                                })
                                auth_conn.commit()
                            
                            st.success(f"✅ User '{new_username}' added successfully!")
                            st.rerun()
                        except Exception as e:
                            if "UNIQUE constraint failed" in str(e):
                                st.error(f"❌ Username '{new_username}' already exists!")
                            else:
                                st.error(f"❌ Error adding user: {str(e)}")
                    else:
                        st.error("❌ Username and password are required!")
        
        with col2:
            st.header("📋 Existing Users")
            
            with auth_engine.connect() as auth_conn:
                users_result = auth_conn.execute(text("SELECT * FROM users ORDER BY created_at DESC"))
                users = users_result.mappings().fetchall()
            
            if users:
                for user in users:
                    with st.expander(f"👤 {user['username']} ({user['role']})", expanded=False):
                        st.write(f"**Full Name:** {user['name'] or 'Not provided'}")
                        st.write(f"**Email:** {user['email'] or 'Not provided'}")
                        st.write(f"**Contact:** {user['contact_number'] or 'Not provided'}")
                        st.write(f"**Role:** {user['role']}")
                        st.write(f"**Created:** {user['created_at']}")
                        
                        # Prevent admin from deleting themselves
                        if user['username'] != st.session_state['username']:
                            if st.button(f"🗑 Delete User", key=f"delete_user_{user['id']}"):
                                try:
                                    with auth_engine.connect() as auth_conn:
                                        auth_conn.execute(text("DELETE FROM users WHERE id = :id"), {'id': user['id']})
                                        auth_conn.commit()
                                    st.success(f"✅ User '{user['username']}' deleted!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Error deleting user: {str(e)}")
                        else:
                            st.info("ℹ️ Cannot delete your own account")
            else:
                st.info("ℹ No users found.")
        
        # User statistics
        st.subheader("📊 User Statistics")
        total_users = len(users) if users else 0
        admin_users = len([u for u in users if u['role'] == 'admin']) if users else 0
        staff_users = len([u for u in users if u['role'] == 'staff']) if users else 0
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Users", total_users)
        with col2:
            st.metric("Admin Users", admin_users)
        with col3:
            st.metric("Staff Users", staff_users)

    # Access denied for non-admin trying to access admin features
    elif selected_tab == "👤 User Management" and st.session_state['role'] != 'admin':
        st.error("❌ Access Denied: Admin privileges required for User Management")

# Main execution logic
if __name__ == "__main__":
    # Check if user is logged in
    if not st.session_state['logged_in']:
        login_page()
    else:
        try:
            main_app()
        except Exception as e:
            st.error(f"❌ Application Error: {str(e)}")
            st.info("Please refresh the page or contact the administrator.")
        finally:
            # Close database connections
            if 'cursor' in locals():
                cursor.close()
            if 'conn' in locals():
                conn.close()